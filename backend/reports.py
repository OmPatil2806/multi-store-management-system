from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from models import Product

COLUMNS = [
    "Name",
    "Category",
    "SKU",
    "Price",
    "Cost Price",
    "Quantity",
    "Low Stock Threshold",
    "Status",
]

COLUMN_WIDTHS = {
    "Name": 28,
    "Category": 16,
    "SKU": 14,
    "Price": 10,
    "Cost Price": 12,
    "Quantity": 10,
    "Low Stock Threshold": 18,
    "Status": 12,
}

HEADER_FONT = Font(bold=True)
LOW_STOCK_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")


def _product_rows(products: list[Product]) -> list[dict]:
    rows = []
    for p in products:
        rows.append(
            {
                "Name": p.name,
                "Category": p.category or "",
                "SKU": p.sku or "",
                "Price": float(p.price),
                "Cost Price": float(p.cost_price) if p.cost_price is not None else None,
                "Quantity": p.quantity,
                "Low Stock Threshold": p.low_stock_threshold,
                "Status": "Low Stock" if p.quantity <= p.low_stock_threshold else "OK",
            }
        )
    return rows


def _write_sheet(writer: pd.ExcelWriter, sheet_name: str, products: list[Product]) -> None:
    # Excel sheet names: max 31 chars, no \ / ? * [ ] : — store names are
    # short enough in practice, but truncate defensively.
    safe_name = sheet_name[:31]
    df = pd.DataFrame(_product_rows(products), columns=COLUMNS)
    df.to_excel(writer, sheet_name=safe_name, index=False)

    worksheet = writer.sheets[safe_name]

    for col_idx in range(1, len(COLUMNS) + 1):
        worksheet.cell(row=1, column=col_idx).font = HEADER_FONT
        worksheet.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[COLUMNS[col_idx - 1]]

    status_col = COLUMNS.index("Status") + 1
    for row_idx in range(2, len(products) + 2):
        if worksheet.cell(row=row_idx, column=status_col).value == "Low Stock":
            for col_idx in range(1, len(COLUMNS) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = LOW_STOCK_FILL


def generate_stock_report(store_products: dict[str, list[Product]]) -> BytesIO:
    """One sheet per store, named after the store. store_products is an
    ordered {store_name: [Product, ...]} mapping — pass a single entry for a
    one-store report or all 3 for the owner's combined export."""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for store_name, products in store_products.items():
            _write_sheet(writer, store_name, products)
    buffer.seek(0)
    return buffer
